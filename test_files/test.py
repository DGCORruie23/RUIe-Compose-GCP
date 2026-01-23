import os
from django.db.models import Count
from usuario.models import Usuario, Paises, EstadoFuerza, Frases, Municipios, PuntosInternacion, RescatePunto, ConteoRapidoPunto, MsgUpdate
from datetime import date
from datetime import *
from openpyxl import Workbook

#----------------------------------------------------
# valores_duplicados = RescatePunto.objects.all() \
#         .values('nombre', 'apellidos', 'nacionalidad') \
#         .annotate(veces=Count('idRescate')) \
#         .filter(veces__gt=1) \
#         .order_by('-veces')

# rescates_totales = RescatePunto.objects.all()

# print(f"Rescates totales: {rescates_totales.count()}")
# print(f"Rescates totales: {valores_duplicados.count()}")

# # ---- Embarazadas
# rescates_totales = RescatePunto.objects.all()

# embarazadas_ORs = RescatePunto.objects.filter(embarazo=True) \
#             .values('oficinaRepre') \
#             .annotate(total=Count('idRescate')) \
#             .order_by('-total')

# embarazadas_Nac = RescatePunto.objects.filter(embarazo=True) \
#             .values('nacionalidad') \
#             .annotate(total=Count('idRescate')) \
#             .order_by('-total')

# print("ORS")
# for dato in embarazadas_ORs:
#     print(f'{dato["oficinaRepre"]},{dato["total"]}')

# print("Nacionalidades")
# for dato in embarazadas_Nac:
#     print(f'{dato["nacionalidad"]},{dato["total"]}')

# # ------ termina embarazadas
#--------------------------------------------------------------



# --------------------------------------------------------------------------------
# --------------------------------------------------------------------------------
# ------------------ reincidentes por fecha --------------

# # # Datos por fechas

# fecha_inicio = date(2025, 1, 1)
# fecha_fin = date(2025, 12, 31)

# # oficinaR = 'BAJA CALIFORNIA'

# fechaIN = datetime.strptime(f"{fecha_inicio}", "%Y-%m-%d")
# fechaFN = datetime.strptime(f"{fecha_fin}", "%Y-%m-%d")

# array_fechas = [(fechaIN + timedelta(days=d)).strftime("%d-%m-%y") for d in range((fechaFN - fechaIN).days + 1)]

# # rescates_por_oficina = RescatePunto.objects.filter(fecha__in= array_fechas, oficinaRepre=oficinaR) \
# rescates_por_oficina = RescatePunto.objects.filter(fecha__in= array_fechas) \
#     .values('oficinaRepre', 'puntoEstra')\
#     .annotate(veces=Count('idRescate')) \
#     .order_by('oficinaRepre')

# # counts_by_date = { entry['fecha']: entry['veces'] for entry in rescates_por_oficina }

# # # 3) construir la lista final incluyendo ceros
# # resultados = []

# # for d in array_fechas:
# #     resultados.append({
# #         'fecha': d,
# #         'veces': counts_by_date.get(d, 0)
# #     })

# # # rescates_por_agente = RescatePunto.objects.filter(fecha__in= array_fechas) \
# # #     .values('oficinaRepre', 'nombreAgente') \
# # #     .annotate(veces=Count('idRescate')) \
# # #     .order_by('oficinaRepre', 'nombreAgente')


# # Crear el archivo Excel
# ruta_archivo_excel = os.path.join(os.getcwd(), f'test_files/Resc_puntos_{fecha_inicio.day:02}_{fecha_inicio.month:02}.xlsx')
# wb = Workbook()
# ws = wb.active
# ws.title = "Rescates por Punto"

# # Escribir la cabecera
# ws.append(["oficinaRepre", "puntoEstra", "Total"])

# # Llenar el archivo Excel con los datos
# for registro in rescates_por_oficina:
#     ws.append([registro['oficinaRepre'], registro['puntoEstra'], registro['veces']])

# # Guardar el archivo
# wb.save(ruta_archivo_excel)
# print(f"El archivo Excel ha sido guardado en {ruta_archivo_excel}")


# --------------------------------------------------------------------------------
# --------------------------------------------------------------------------------


# --------------------------------------------------------------------------------
# --------------------------------------------------------------------------------
# ------------------ Funcion para cambiar de minusculas a mayusculas --------------

# from usuario.models import Usuario, Paises, EstadoFuerza, Frases, Municipios, PuntosInternacion, RescatePunto
# from django.db.models.functions import Upper
# RescatePunto.objects.update(nacionalidad=Upper('nacionalidad'))
# RescatePunto.objects.update(nombre=Upper('nombre'))
# RescatePunto.objects.update(apellidos=Upper('apellidos'))
# RescatePunto.objects.update(nombreAgente=Upper('nombreAgente'))
# RescatePunto.objects.update(puntoEstra=Upper('puntoEstra'))

# ------------------------------------------------------------------
# ------------------------------------------------------------------


# ------------------------------------------------------------------
# ------------------------------------------------------------------
# ----------  Descargar datos por año --------------------

def exportar_excel(fecha_inicio, fecha_fin):

    LIMITE_POR_HOJA = 500_000
    
    ruta_archivo_excel = os.path.join(os.getcwd(), f'test_files/Rescates_{fecha_inicio.day:02}_{fecha_inicio.month:02}_{fecha_inicio.year:02}__{fecha_fin.day:02}_{fecha_fin.month:02}_{fecha_fin.year:02}.xlsx')
    
    fechaIN = datetime.strptime(f"{fecha_inicio}", "%Y-%m-%d")
    fechaFN = datetime.strptime(f"{fecha_fin}", "%Y-%m-%d")

    array_fechas = [(fechaIN + timedelta(days=d)).strftime("%d-%m-%y") for d in range((fechaFN - fechaIN).days + 1)]

    wb = Workbook()
    ws = wb.active
    ws.title = "Hoja_1"

    encabezados = ['oficinaRepre', 'fecha', 'hora', 'casaSeguridad', 'voluntarios', 'puestosADispo', 'puntoEstra', 'nacionalidad', 'nombre', 'apellidos', 'fechaNacimiento', 'sexo', 'edad', 'numFamilia', 'embarazo']
    ws.append(encabezados)

    fila_actual = 1
    hoja_num = 1

    queryset = (
        RescatePunto.objects.filter(fecha__in= array_fechas)
        .values('oficinaRepre', 'fecha', 'hora', 'casaSeguridad', 'voluntarios', 'puestosADispo', 'puntoEstra', 'nacionalidad', 'nombre', 'apellidos', 'fechaNacimiento', 'sexo', 'edad', 'numFamilia', 'embarazo')\
        .order_by('oficinaRepre')
        .iterator(chunk_size=10_000)
    )

    for obj in queryset:
        # Si llegamos al límite → nueva hoja
        if fila_actual >= LIMITE_POR_HOJA:
            hoja_num += 1
            ws = wb.create_sheet(title=f"Hoja_{hoja_num}")
            ws.append(encabezados)
            fila_actual = 1

        ws.append([
            obj['oficinaRepre'],
            obj['fecha'],
            obj['hora'],
            obj['casaSeguridad'],
            obj['voluntarios'],
            obj['puestosADispo'],
            obj['puntoEstra'],
            obj['nacionalidad'],
            obj['nombre'],
            obj['apellidos'],
            obj['fechaNacimiento'],
            obj['sexo'],
            obj['edad'],
            obj['numFamilia'],
            obj['embarazo'],
        ])

        fila_actual += 1

    # Guardar el archivo
    wb.save(ruta_archivo_excel)
    print(f"El archivo Excel ha sido guardado en {ruta_archivo_excel}")

exportar_excel(fecha_inicio = date(2024, 1, 1), fecha_fin = date(2024, 12, 31))